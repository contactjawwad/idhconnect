#jawwad@jumphost:~/InventoryDataHub/app/controllers/sfp_controller.py

from flask import Blueprint,session,render_template,current_app,request,jsonify,send_file,Response
from .main_controller import BaseController
from ..services.sfp_service import SFPService  
import re
import tempfile
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import  PatternFill, Font, Alignment, Border, Side


sfp_blueprint=Blueprint('sfp',__name__)
#This sfp used in argument is used in HTML sfp.

@sfp_blueprint.route('/sfp',methods=['GET'])
def sfp_report_route():
    print("sfp_report_route called.")
    controller=SFPController()
    uploaded_files = session.get('temp_files', [])  # Get data from session
    temp_dir = current_app.config['temp_dir']
    return controller.sfp_report(uploaded_files,temp_dir)  # Pass the data to the controller method
    # THE ABOVE CODE IS NOT REQUIRED AS NOW IT WILL BE COVERED IN THE sfp_report_data 


@sfp_blueprint.route('/sfp/data', methods=['GET'])
def sfp_report_data():    
    controller = SFPController()
    uploaded_files = session.get('temp_files', [])
    temp_dir = current_app.config['temp_dir']
    start = int(request.args.get('start', 0))
    chunk_size = int(request.args.get('length', 10000))
    print(f"Controller Requesting chunk: start={start}, length={chunk_size}")
    return controller.sfp_report_data(uploaded_files, temp_dir, start, chunk_size)

@sfp_blueprint.route('/sfp/export', methods=['GET'])
def sfp_export_route():
    controller = SFPController()
    return controller.export_sfp_report()


class SFPController(BaseController):

    def __init__(self):
        #Pass Template Name and Title Name , which can be used in rendering
        super().__init__('sfp_report.html', 'SFP Report')
        self.service=SFPService()
        print("SFP Controller Initialized")

    def sfp_report(self,uploaded_files:list[str],temp_dir:str):
        
        #table_data, summary_data = self.service.process_files(uploaded_files, temp_dir) 
        #return self.render({'sfp_files': table_data, 'summary_data': summary_data})
        return self.render()
    
    def sfp_report_data(self, uploaded_files: list[str], temp_dir: str, start: int, chunk_size: int):
        table_data, summary_data, all_data_fetched = self.service.process_files(uploaded_files, temp_dir, start, chunk_size)
        print(f"Controller Class Returning chunk to JS: {len(table_data)} rows, start={start}, length={chunk_size}")
        
        return jsonify({'data': table_data, 'summary_data': summary_data, 'all_data_fetched': all_data_fetched})
        
    def export_sfp_report(self):
        try:
            # 1) Fetch data from session and temp directory
            uploaded_files = session.get('temp_files', [])
            temp_dir        = current_app.config['temp_dir']

            # 2) Process files in 10k-row chunks
            table_data, summary_data, _ = self.service.process_files(
                uploaded_files, temp_dir, start=0, chunk_size=10**4
            )

            # 3) Create a standard Workbook (no write_only) for styling
            wb = Workbook()
            ws_main = wb.active
            ws_main.title = 'Main Report'

            # 4) Define styles
            header_fill      = PatternFill('solid', fgColor='FF000080')  # dark blue
            header_font      = Font(bold=True, color='FFFFFFFF')         # white bold
            header_align     = Alignment(horizontal='center', vertical='center')
            row_fill         = PatternFill('solid', fgColor='FFDCE6F1')  # light blue
            thin_side        = Side(style='thin', color='000000')
            thin_border      = Border(top=thin_side, left=thin_side, bottom=thin_side, right=thin_side)

            # 5) Add and style header row
            headers = ['Site Name', 'Connector Type', 'Part Number', 'Vendor Serial Number', 'Description', 'Shelf Type']
            ws_main.append(headers)
            for cell in ws_main[1]:  # first row
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = header_align
                cell.border    = thin_border

            # 6) Append data rows with styling
            for row in table_data:
                ws_main.append([
                    row.get('Site Name'),
                    row.get('Connector Type'),
                    row.get('Part Number'),
                    row.get('Vendor Serial Number'),
                    row.get('Description'),
                    row.get('Shelf Type', '')
                ])
                for cell in ws_main[ws_main.max_row]:
                    cell.fill      = row_fill
                    cell.alignment = header_align
                    cell.border    = thin_border

            # 7) Auto-size columns
            for column in ws_main.columns:
                max_len = max(len(str(cell.value or '')) for cell in column)
                ws_main.column_dimensions[column[0].column_letter].width = max_len + 2

            # 8) Create and style Summary Report sheet
            ws_sum = wb.create_sheet('Summary Report')
            ws_sum.append(['Part Number', 'QTY', 'Description'])
            for cell in ws_sum[1]:
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = header_align
                cell.border    = thin_border
            for pn, info in summary_data.items():
                ws_sum.append([pn, info['QTY'], info['Description']])
                for cell in ws_sum[ws_sum.max_row]:
                    cell.fill      = row_fill
                    cell.alignment = header_align
                    cell.border    = thin_border

            for column in ws_sum.columns:
                max_len = max(len(str(cell.value or '')) for cell in column)
                ws_sum.column_dimensions[column[0].column_letter].width = max_len + 2

            # 9) Stream the Excel file to client
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return send_file(
                output,
                as_attachment=True,
                download_name='SFP_Model_Report.xlsx',
                mimetype=(
                    'application/vnd.openxmlformats-officedocument.'
                    'spreadsheetml.sheet'
                )
            )

        except Exception:
            current_app.logger.exception('Failed to export SFP report')
            return jsonify({'error': 'Export failed on server'}), 500
